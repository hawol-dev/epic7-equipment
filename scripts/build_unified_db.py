"""
통합 영웅 DB + enums + 검색 인덱스 생성 (영문 ID 정규화 버전)

소스:
- data/processed/heroes_merged.json    (파일1+파일2)
- data/raw/fribbels_herodata.json      (영문/이미지/스킬)
- data/raw/fribbels_ko.json            (한글번역)

출력:
- data/processed/heroes.json           (정규화된 영웅 DB)
- data/processed/enums.json            (모든 enum 정의 + 라벨)
- data/processed/search_index.json     (영문 ID 기반 인덱스)
- data/processed/build_report.txt

영웅 type (Fribbels code prefix 기반):
- standard:       c1xxx    일반 성약 4~5★
- moonlight5:     c2xxx 5★ 월광 (Light/Dark 얼터)
- moonlight4:     c2xxx 4★ 월광
- moonlight4_alt: c6xxx 4★ 월광/한정
- regular3:       c3xxx    일반 3★
- specialty3:     c4xxx    전직 3★
- limited:        c5xxx    한정/콜라보/스토리
- unknown:        m코드 등 특수
"""
import json
import sys
import io
import re
import openpyxl
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent))
from enums import (
    SUBSTATS, SETS, ELEMENTS, CLASSES, ZODIACS, ENGRAVINGS,
    SET_KO_TO_ID, SUBSTAT_KO_TO_ID,
    ELEMENT_KO_TO_ID, CLASS_KO_TO_ID, ZODIAC_KO_TO_ID, ENGRAVING_KO_TO_ID,
    FRIBBELS_ELEMENT_TO_ID, FRIBBELS_CLASS_TO_ID, FRIBBELS_ZODIAC_TO_ID,
    export_enums_json,
)
from hero_guides import HERO_GUIDES
from hero_notes_en import NOTE_EN

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(r"E:\jsh02\Dev\EpicSevenEquipment")
SRC_MERGED = ROOT / "data/processed/heroes_merged.json"
SRC_FRIBBELS = ROOT / "data/raw/fribbels_herodata.json"
SRC_FRIBBELS_KO = ROOT / "data/raw/fribbels_ko.json"
SRC_FILE2 = Path(r"C:\Users\jsh02\Downloads\에픽세븐장비시뮬_26_04_14의 사본.xlsx")

DST_HEROES = ROOT / "data/processed/heroes.json"
DST_ENUMS = ROOT / "data/processed/enums.json"
DST_INDEX = ROOT / "data/processed/search_index.json"
DST_REPORT = ROOT / "data/processed/build_report.txt"

# 한글번역 누락 영웅 매뉴얼 매핑 (Fribbels에 한글 번역 없는 경우 채움)
MANUAL_EN_KO_NAMES = {
    "Hecate": "헤카테", "Notos": "노토스", "Salome": "살로메",
    "Serila": "세리라", "Aki": "아키", "Fern": "페른",
    "Frieren": "프리렌", "Stark": "슈타르크", "Saria": "사르미아",
    "Ruiza": "루이자", "Wretched Rose": "비탄의 로제",
    "School Nurse Yulha": "보건교사 율하",
    "Sealed Eye Surin": "봉안의 수린",
    "Shepherd of the Dark Diene": "어둠의 목자 디에네",
    "Monarch of the Sword Iseria": "보검의 군주 이세리아",
    "Tactical Archetype Coli": "전술형 콜리",
    "Perfumer Byblis": "조향사 비브리스",
    "Estelle": "에스텔",
}

# Fribbels 한글 번역이 잘못된 경우 강제 교정 (정식 게임 명칭으로)
KO_NAME_OVERRIDES = {
    "Bomb Model Kanna": "폭격형 카논",         # ko_dict는 "카논"으로 약식 표기
    "Sigret": "세크레트",                       # ko_dict는 "세크레트" 맞음 (확인용)
    "Operator Sigret": "오퍼레이터 세크레트",   # ko_dict는 "오퍼레이터 세크레트" 맞음
    "Swift Flagbearer Sigret": "쾌속의 기수 세크레트",  # 이전엔 "시그렛"으로 잘못 매핑
}

# file1의 base_name 표기를 정식 명칭으로 교정 (alias)
# file1엔 약식/구표기로 들어있으나 정식은 다른 경우
FILE1_BASE_ALIAS = {
    "적월의 헤이스트": "적월의 귀족 헤이스트",   # 정식: 적월의 귀족 헤이스트
    "쾌속의 기수 세크레트": "쾌속의 기수 세크레트",  # 그대로 (KO_NAME_OVERRIDES 후 매칭됨)
    "폭격형 카논": "폭격형 카논",               # 그대로 (KO_NAME_OVERRIDES 후 매칭됨)
    "전도사 카마인로즈": "전도자 카마인로즈",     # 사 vs 자 오타
}

# DB에서 완전히 제외할 영웅들 (Fribbels EN 이름 기준)
# 강화판으로 대체되는 베이스 영웅 등
EXCLUDED_FRIBBELS = {
    "Ras",  # 모험가 라스(c5001) 강화판으로 대체
}

# file1에 유효옵 정보는 없지만 사이트에는 노출할 영웅 (Fribbels EN 이름)
# 유효옵 데이터 추가되기 전까진 has_data=false로 표시됨
INCLUDE_FRIBBELS_ONLY = {
    "Vivian",  # 비비안 — 자연 마도사 5★ 베이스 (사용자 요청)
}

# file1 variant 한글 표기 → URL 안전한 영문 슬러그
# (영웅 ID는 ASCII만 — URL 인코딩 이슈 방지)
VARIANT_SLUG = {
    "PVE": "pve",
    "PVP": "pvp",
    "성좌": "ext",
    "성좌용": "ext",
    "원정대": "exp",
    "원정대 리치 전열": "exp-lich",
    "딜": "dps",
    "탱": "tank",
    "딜탱": "dpstank",
    "와이번": "wyvern",
    "밴시원펀": "banshee-1p",
    "골렘원펀": "golem-1p",
    "속덱": "spd-deck",
    "선턴": "first-turn",
    "중턴": "mid-turn",
    "저항": "res",
    "속막이": "spd-block",
    "흡혈": "lifesteal",
    "반혈랑": "halfhp",
    "효저": "effres",
    "치치": "crit",
    "대 스트라제스용": "vs-strazes",
    "스탠다드": "std",
    "전당": "wos",
    "수호": "guard",
    "적": "eff",
    "적저": "eff-res",
    "효저단": "effres-only",
    "아레나공덱": "arena-atk",
    "효적": "eff-acc",
    "반교집": "anti-cheese",
    "디버퍼": "debuffer",
    "속": "spd",
    "하르세티연계": "vs-harsetti",
}


def load_recommended_artifacts() -> dict[str, list[str]]:
    """
    파일2 '추천 세팅' 시트 → {hero_ko_name: [artifact1, artifact2, artifact3]}
    빈 칸은 제외, 한 개도 없으면 영웅 자체를 dict에서 제외.
    """
    wb = openpyxl.load_workbook(SRC_FILE2, data_only=True, read_only=True)
    ws = wb["추천 세팅"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # 0=헤더, 1=영웅 선택 placeholder
            continue
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        if name == "영웅 선택":
            continue
        artifacts = [row[2], row[3], row[4]]
        artifacts = [str(a).strip() for a in artifacts if a]
        if artifacts:
            out[name] = artifacts
    wb.close()
    return out


def load_engraving_grades() -> dict[str, dict[int, dict[str, str | None]]]:
    """
    파일2 '계산요소' 시트 → 각인집중 등급별 효과 표
    {stat_id: {level(3/4/5): {grade(D~SSS): "21%" or None}}}
    stat_id는 ENGRAVING_KO_TO_ID로 변환된 값 사용
    """
    GRADES = ["D", "C", "B", "A", "S", "SS", "SSS"]
    wb = openpyxl.load_workbook(SRC_FILE2, data_only=True, read_only=True)
    ws = wb["계산요소"]
    out: dict = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        ko_stat = row[0]
        level = row[1]
        if not ko_stat or level is None:
            continue
        ko_stat = str(ko_stat).strip()
        try:
            level_int = int(level)
        except (TypeError, ValueError):
            continue
        # ENGRAVING_KO_TO_ID는 "공격력(%)" 같은 표기 사용
        sid = ENGRAVING_KO_TO_ID.get(ko_stat)
        if not sid:
            continue
        # cols 3..9 = D..SSS
        grade_values = {}
        for gi, grade in enumerate(GRADES):
            v = row[2 + gi]
            grade_values[grade] = str(v).strip() if v is not None else None
        out.setdefault(sid, {})[level_int] = grade_values
    wb.close()
    return out


# Fribbels CDN URL → 로컬 이미지 경로 (web/public 기준)
def local_image_path(url: str | None) -> str | None:
    if not url:
        return None
    # 마지막 파일명만 추출: .../cachedimages/c1019_s.png → c1019_s.png
    filename = url.rsplit("/", 1)[-1]
    if not filename:
        return None
    return f"/images/heroes/{filename}"


# Fribbels code prefix는 c5xxx인데 실제로는 한정이 아니라 성약 영웅인 케이스
# (Fribbels의 분류와 한국 게임 내 분류 차이 보정)
COVENANT_C5_OVERRIDES = {
    "c5001",  # 모험가 라스
    "c5033",  # 전술형 콜리
}


# 영웅 type + code → 카테고리 태그 (필터 그룹용)
# 한 영웅이 여러 카테고리에 속할 수 있음
def categories_from_type(t: str, code: str | None = None) -> list[str]:
    if t == "limited" and code in COVENANT_C5_OVERRIDES:
        return ["covenant"]  # 한정 아닌 일반 성약
    if t == "standard":
        return ["covenant"]
    if t == "limited":
        return ["covenant", "limited"]
    if t.startswith("moonlight"):
        return ["moonlight"]
    if t == "specialty3":
        return ["specialty"]
    if t == "regular3":
        return ["regular3"]
    return []


# Fribbels code prefix → 영웅 type
def classify_type(code: str, rarity) -> str:
    if not code:
        return "unknown"
    m = re.match(r"^c(\d)(\d+)", code)
    if not m:
        return "unknown"
    prefix = m.group(1)
    if prefix == "1":
        return "standard"
    if prefix == "2":
        if rarity == 5:
            return "moonlight5"
        if rarity == 4:
            return "moonlight4"
        return "moonlight"
    if prefix == "3":
        return "regular3"
    if prefix == "4":
        return "specialty3"
    if prefix == "5":
        return "limited"
    if prefix == "6":
        return "moonlight4_alt"
    return "unknown"


def slugify_kr(s: str) -> str:
    """기본 slug — 한글 그대로 유지 (Korean fallback용, ASCII 보장 X)"""
    s = s.strip()
    s = re.sub(r"\s*\(", "__", s)
    s = re.sub(r"\)\s*", "", s)
    s = re.sub(r"\s+", "_", s)
    return s


def variant_slug(variant_ko: str | None, used_in_base: set[str], base_id: str) -> str | None:
    """variant 한글 → ASCII 슬러그. 알려진 매핑 우선, 모르면 v1/v2/... 인덱스"""
    if not variant_ko:
        return None
    slug = VARIANT_SLUG.get(variant_ko)
    if slug:
        return slug
    # 매핑 없으면 base 별로 v1, v2 인덱스
    n = sum(1 for x in used_in_base if x.startswith(f"{base_id}__v")) + 1
    return f"v{n}"


def main():
    print("Loading sources...")
    heroes_ko = json.loads(SRC_MERGED.read_text(encoding="utf-8"))
    fribbels = json.loads(SRC_FRIBBELS.read_text(encoding="utf-8"))
    ko_dict = json.loads(SRC_FRIBBELS_KO.read_text(encoding="utf-8"))

    # 추가 데이터 (파일2)
    artifacts = load_recommended_artifacts()  # {ko_name: [art1, art2, art3]}
    engraving_grades = load_engraving_grades()  # {sid: {lv: {grade: val}}}
    print(f"  추천 아티팩트: {len(artifacts)}명, 각인 등급표: {len(engraving_grades)} 스탯")

    # 아티팩트 ko → en 매핑 (Fribbels artifactdata.json + ko_dict 역매핑)
    try:
        art_en_to_ko = {}
        # ko_dict 안에 아티팩트 영문 키들도 들어있음 — Fribbels artifactdata에서 영문 키 가져와서 ko_dict에서 한글 찾기
        from urllib.request import urlopen
        # 캐시: 한 번 다운받아서 raw에 저장돼있을 가능성 — local file 우선
        art_local = ROOT / "data/raw/fribbels_artifacts.json"
        if not art_local.exists():
            art_data = json.loads(urlopen(
                "https://raw.githubusercontent.com/fribbels/Fribbels-Epic-7-Optimizer/main/data/cache/artifactdata.json"
            ).read())
            art_local.write_text(json.dumps(art_data, ensure_ascii=False), encoding="utf-8")
        else:
            art_data = json.loads(art_local.read_text(encoding="utf-8"))
        # 영문 키들 → ko_dict로 한글 매핑
        artifact_ko_to_en = {}
        for en_name in art_data.keys():
            ko = ko_dict.get(en_name)
            if ko:
                artifact_ko_to_en[ko] = en_name
        # Fribbels 번역에 없는 신규 아티팩트 수동 매핑
        MANUAL_ARTIFACTS = {
            "금강 각반": "Diamond Greaves",
            "소중한 인연": "A Precious Connection",
            "윈드라이너": "Windrider",
            "푸른 장미의 가시": "Thorns of the Blue Rose",
            "하사받은 펜": "Bestowed Pen",
            "경련화 반지": "Spasm Ring",
            "만능 회복 요술봉": "All-Purpose Recovery Wand",
            "눈을 뜬 잎새": "Awakened Leaf",
            "에티가 셉터": "Etica Scepter",
            "염원의 일격": "Strike of Aspiration",
            "악몽의 주인": "Lord of Nightmares",
            "해방된 참전의 도끼": "Liberated Axe of War",
            "나비 머리핀": "Butterfly Hairpin",
            "봉염의 의식": "Sealing Flame Ritual",
            "진실의 향수": "Perfume of Truth",
            "쉐도우 윈즈7": "Shadow Wings 7",
            "여신의 검은손": "Black Hand of the Goddess",
            "영광의 깃발": "Banner of Glory",
            "M.O.A.S": "M.O.A.S",
            "수명의 서": "Book of Life",
        }
        for ko, en in MANUAL_ARTIFACTS.items():
            artifact_ko_to_en.setdefault(ko, en)
        print(f"  아티팩트 ko↔en 매핑: {len(artifact_ko_to_en)} ({len(MANUAL_ARTIFACTS)} 수동)")
    except Exception as e:
        print(f"  ⚠ 아티팩트 매핑 실패: {e}")
        artifact_ko_to_en = {}

    # EN → KO 매핑 (우선순위: 강제 override > Fribbels ko_dict > 수동 매핑)
    en_to_ko = {}
    ko_to_en = {}
    for en_name in fribbels.keys():
        ko_name = (
            KO_NAME_OVERRIDES.get(en_name)
            or ko_dict.get(en_name)
            or MANUAL_EN_KO_NAMES.get(en_name)
        )
        if ko_name:
            en_to_ko[en_name] = ko_name
            ko_to_en[ko_name] = en_name

    print(f"  파일1+2: {len(heroes_ko)}, Fribbels: {len(fribbels)}, EN↔KO: {len(en_to_ko)}")

    out = []
    seen_ids = set()
    fribbels_used = set()
    warnings = []

    def norm_substats(ko_substats):
        result = {}
        for ko_key, level in ko_substats.items():
            sid = SUBSTAT_KO_TO_ID.get(ko_key)
            if not sid:
                warnings.append(f"unknown substat KO: '{ko_key}'")
                continue
            result[sid] = level
        # 8개 모두 보장
        for sid in SUBSTATS:
            result.setdefault(sid, None)
        return result

    def norm_priority_order(ko_order):
        result = []
        for ko_key in ko_order:
            sid = SUBSTAT_KO_TO_ID.get(ko_key)
            if sid:
                result.append(sid)
            else:
                warnings.append(f"unknown priority KO: '{ko_key}'")
        return result

    def norm_set_list(ko_sets, ctx=""):
        result = []
        for ko_set in ko_sets:
            sid = SET_KO_TO_ID.get(ko_set)
            if sid:
                result.append(sid)
            else:
                warnings.append(f"unknown set KO: '{ko_set}' ({ctx})")
        return result

    def norm_set_combos(alternates):
        return [norm_set_list(alt, "set_combo") for alt in alternates]

    def norm_engraving(eng_ko):
        if not eng_ko:
            return None
        eid = ENGRAVING_KO_TO_ID.get(eng_ko)
        if not eid:
            warnings.append(f"unknown engraving KO: '{eng_ko}'")
        return eid

    # 파일1 영웅 처리
    for h in heroes_ko:
        ko_name = h["name"]
        base_ko = h["base_name"]
        variant_ko = h["variant"]

        # alias로 정식 명칭 변환 후 매칭 시도
        lookup_base = FILE1_BASE_ALIAS.get(base_ko, base_ko)
        en_name = ko_to_en.get(lookup_base)
        fr = fribbels.get(en_name) if en_name else None
        if fr:
            fribbels_used.add(en_name)

        meta = h.get("meta") or {}
        rarity = meta.get("rank")
        element = CLASS_KO_TO_ID.get(meta.get("class") or "")  # placeholder, fixed below

        # element / class / zodiac 결정 (file2 메타 우선, fribbels 보조)
        element = ELEMENT_KO_TO_ID.get(meta.get("element") or "")
        if not element and fr:
            element = FRIBBELS_ELEMENT_TO_ID.get(fr.get("attribute"))

        char_class = CLASS_KO_TO_ID.get(meta.get("class") or "")
        if not char_class and fr:
            char_class = FRIBBELS_CLASS_TO_ID.get(fr.get("role"))

        zodiac = ZODIAC_KO_TO_ID.get(meta.get("zodiac") or "")
        if not zodiac and fr:
            zodiac = FRIBBELS_ZODIAC_TO_ID.get(fr.get("zodiac"))

        if rarity is None and fr:
            rarity = fr.get("rarity")

        engraving = norm_engraving(meta.get("engraving_focus"))

        image = None
        if fr:
            assets = fr.get("assets", {})
            image = {
                "icon": local_image_path(assets.get("icon")),
                "thumbnail": local_image_path(assets.get("thumbnail")),
            }

        # ID 생성 (영문 우선, ASCII 보장)
        if en_name and fr:
            base_id = fr.get("_id") or en_name.lower().replace(" ", "_")
        else:
            # Fribbels 매칭 실패 — 한글 fallback (URL 인코딩 됨)
            base_id = slugify_kr(base_ko).lower()
        if variant_ko:
            v_slug = variant_slug(variant_ko, seen_ids, base_id)
            full_id = f"{base_id}__{v_slug}"
        else:
            full_id = base_id
        if full_id in seen_ids:
            full_id = f"{full_id}_{len(out)}"
        seen_ids.add(full_id)

        sources = ["file1"]
        if h.get("meta"):
            sources.append("file2")
        if fr:
            sources.append("fribbels")

        notes_ko = h["notes"]
        notes_en = NOTE_EN.get(notes_ko) if notes_ko else None
        valid_options = {
            "substats": norm_substats(h["valid_substats"]),
            "priority_order": norm_priority_order(h["priority"]["order"]),
            "priority_unknown": h["priority"]["unknown"],
            "set_combos": norm_set_combos(h["set_combo"]["alternates"]),
            "valid_sets": norm_set_list(h["valid_sets"], f"valid_sets({ko_name})"),
            "ignore_2set": h["set_combo"]["ignore_2set"],
            "notes": notes_ko,
            "notes_en": notes_en,
        }

        code = fr.get("code") if fr else None

        # 기본 스탯 (Fribbels lv50/lv60 5★ Fully Awakened)
        base_stats = None
        if fr:
            cs = fr.get("calculatedStatus") or {}
            base_stats = {
                "lv50_5": cs.get("lv50FiveStarFullyAwakened"),
                "lv60_6": cs.get("lv60SixStarFullyAwakened"),
            }

        # 추천 아티팩트 (file1 영웅 base_name 기준)
        rec_artifacts_ko = artifacts.get(base_ko) or artifacts.get(lookup_base)
        rec_artifacts = None
        if rec_artifacts_ko:
            rec_artifacts = [
                {"ko": a, "en": artifact_ko_to_en.get(a)}
                for a in rec_artifacts_ko
            ]

        # 외부 가이드 링크 (디시 갤러리 등)
        guides = HERO_GUIDES.get(base_ko) or HERO_GUIDES.get(lookup_base)

        type_id = classify_type(code, rarity) if code else "unknown"
        record = {
            "id": full_id,
            "code": code,
            "type": type_id,
            "categories": categories_from_type(type_id, code),
            "names": {"ko": ko_name, "en": en_name},
            "base_name_ko": base_ko,
            "variant_ko": variant_ko,
            "rarity": rarity,
            "element": element,
            "class": char_class,
            "zodiac": zodiac,
            "engraving_focus": engraving,
            "image": image,
            "base_stats": base_stats,
            "recommended_artifacts": rec_artifacts,
            "guides": guides,
            "valid_options": valid_options,
            "has_data": True,
            "source": sources,
        }
        out.append(record)

    # Fribbels-only — 화이트리스트에 있는 영웅만 추가 (나머지는 제외)
    fribbels_only_ids = (set(fribbels.keys())
                         - fribbels_used
                         - EXCLUDED_FRIBBELS) & INCLUDE_FRIBBELS_ONLY
    for en_name in sorted(fribbels_only_ids):
        entry = fribbels[en_name]
        ko_name = en_to_ko.get(en_name)
        display = ko_name or en_name
        base_id = entry.get("_id") or en_name.lower().replace(" ", "_")
        if base_id in seen_ids:
            base_id = f"{base_id}_{len(out)}"
        seen_ids.add(base_id)

        assets = entry.get("assets", {})
        code = entry.get("code")
        rarity = entry.get("rarity")
        type_id = classify_type(code, rarity) if code else "unknown"

        cs = entry.get("calculatedStatus") or {}
        bs = {
            "lv50_5": cs.get("lv50FiveStarFullyAwakened"),
            "lv60_6": cs.get("lv60SixStarFullyAwakened"),
        }

        record = {
            "id": base_id,
            "code": code,
            "type": type_id,
            "categories": categories_from_type(type_id, code),
            "names": {"ko": display, "en": en_name},
            "base_name_ko": display,
            "variant_ko": None,
            "rarity": rarity,
            "element": FRIBBELS_ELEMENT_TO_ID.get(entry.get("attribute")),
            "class": FRIBBELS_CLASS_TO_ID.get(entry.get("role")),
            "zodiac": FRIBBELS_ZODIAC_TO_ID.get(entry.get("zodiac")),
            "engraving_focus": None,
            "image": {
                "icon": local_image_path(assets.get("icon")),
                "thumbnail": local_image_path(assets.get("thumbnail")),
            },
            "base_stats": bs,
            "recommended_artifacts": (
                [{"ko": a, "en": artifact_ko_to_en.get(a)}
                 for a in (artifacts.get(display) or [])]
                if artifacts.get(display)
                else None
            ),
            "valid_options": None,
            "has_data": False,
            "source": ["fribbels"],
        }
        out.append(record)

    # 저장
    DST_HEROES.parent.mkdir(parents=True, exist_ok=True)
    DST_HEROES.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    DST_ENUMS.write_text(
        json.dumps(
            export_enums_json({"engraving_grades": engraving_grades}),
            ensure_ascii=False, indent=2
        ),
        encoding="utf-8"
    )

    # 검색 인덱스 (영문 ID 기반)
    idx = {
        "by_substat_essential": {},  # spd: [hero_ids]
        "by_substat_preferred": {},
        "by_set": {},                 # speed: [hero_ids]
        "by_element": {},
        "by_class": {},
        "by_rarity": {},
        "by_zodiac": {},
        "by_engraving": {},
        "by_type": {},                # standard / moonlight5 / ...
    }
    for r in out:
        if r["element"]:
            idx["by_element"].setdefault(r["element"], []).append(r["id"])
        if r["class"]:
            idx["by_class"].setdefault(r["class"], []).append(r["id"])
        if r["rarity"]:
            idx["by_rarity"].setdefault(str(r["rarity"]), []).append(r["id"])
        if r["zodiac"]:
            idx["by_zodiac"].setdefault(r["zodiac"], []).append(r["id"])
        if r.get("engraving_focus"):
            idx["by_engraving"].setdefault(r["engraving_focus"], []).append(r["id"])
        if r.get("type"):
            idx["by_type"].setdefault(r["type"], []).append(r["id"])
        if r["valid_options"]:
            for sid, level in r["valid_options"]["substats"].items():
                if level == "essential":
                    idx["by_substat_essential"].setdefault(sid, []).append(r["id"])
                elif level == "preferred":
                    idx["by_substat_preferred"].setdefault(sid, []).append(r["id"])
            for set_id in r["valid_options"]["valid_sets"]:
                idx["by_set"].setdefault(set_id, []).append(r["id"])

    DST_INDEX.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")

    # 리포트
    lines = [
        "=== 통합 DB 빌드 리포트 (영문 ID 정규화) ===",
        f"통합 영웅: {len(out)}",
        f"  유효옵 데이터 있음: {sum(1 for r in out if r['has_data'])}",
        f"  유효옵 데이터 없음: {sum(1 for r in out if not r['has_data'])}",
        f"  이미지 있음: {sum(1 for r in out if r.get('image') and r['image'].get('icon'))}",
        f"  한글명 누락: {sum(1 for r in out if r.get('translation_missing'))}",
        "",
        "=== 인덱스 stats ===",
    ]
    for k, d in idx.items():
        lines.append(f"  {k}: {len(d)} keys, {sum(len(v) for v in d.values())} entries")

    if warnings:
        lines.append("")
        lines.append(f"=== 경고 ({len(warnings)}) ===")
        for w in Counter(warnings).most_common(20):
            lines.append(f"  {w[1]}x  {w[0]}")

    DST_REPORT.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines))
    print(f"\n저장:")
    print(f"  {DST_HEROES}")
    print(f"  {DST_ENUMS}")
    print(f"  {DST_INDEX}")


if __name__ == "__main__":
    main()

"""file1에 비비안 관련 영웅이 정확히 어떻게 있는지 확인"""
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jsh02\Downloads\에픽세븐 장비 유효옵의 사본.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["영웅별 유효옵"]

print("=== file1 영웅별 유효옵 시트에서 '비비안' 찾기 ===\n")
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if name and "비비안" in str(name):
        # 전체 행 출력
        row_vals = []
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v) if v is not None else "")
        print(f"행 {r}: {row_vals}")
        print()

print("=== file2 캐릭터 정보에서 '비비안' 찾기 ===\n")
path2 = r"C:\Users\jsh02\Downloads\에픽세븐장비시뮬_26_04_14의 사본.xlsx"
wb2 = openpyxl.load_workbook(path2, data_only=True, read_only=True)
ws2 = wb2["캐릭터 정보"]
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    name = row[0]
    if name and "비비안" in str(name):
        print(f"  '{name}' — 등급:{row[2]}, 속성:{row[3]}, 별:{row[4]}, 직업:{row[5]}")

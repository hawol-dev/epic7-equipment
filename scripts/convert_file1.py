"""
파일1 (영웅별 유효옵.xlsx) → 정규화된 JSON

스키마:
{
  "name": "고독한 늑대 페이라",
  "variant": null,                  # 괄호 안 표기 (예: "PVE", "흡혈")
  "base_name": "고독한 늑대 페이라", # 괄호 제거
  "valid_substats": {
    "속도": "essential",   # ● 필수
    "공격력%": "essential",
    "생명력%": "preferred", # ○ 선호
    "방어력%": null,
    "치명확률": null,
    "치명피해": null,
    "효과적중": "preferred",
    "효과저항": null
  },
  "priority": {
    "raw": "속공생적",
    "order": ["속도", "공격력%", "생명력%", "효과적중"],  # 약자 풀어서
    "unknown": false   # "몰루"이면 true
  },
  "set_combo": {
    "raw": "속추생적면",
    "alternates": [["속도","추격","체력","적중","면역"]]   # / 로 분리
  },
  "valid_sets": ["속도","추격","체력","적중","면역"],  # cols 21-33 의 풀네임
  "notes": "중턴면역 디버퍼적중 선턴은 속도부터"
}
"""
import openpyxl
import json
import sys
import io
import re
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = Path(r"C:\Users\jsh02\Downloads\에픽세븐 장비 유효옵의 사본.xlsx")
DST_RAW = Path(r"E:\jsh02\Dev\EpicSevenEquipment\data\raw\heroes_valid_options.json")
DST_REPORT = Path(r"E:\jsh02\Dev\EpicSevenEquipment\data\raw\conversion_report.txt")

# 컬럼 매핑
COL = {
    "name": 1,
    "speed": 2, "atk": 3, "hp": 4, "def": 5,
    "crit": 6, "critdmg": 7, "eff": 8, "effres": 9,
    "priority": 10,
    "set_combo": 18,
    "notes": 19,
}
SET_KW_COLS = list(range(21, 34))  # 21..33 inclusive

# 부옵 컬럼 헤더와 풀네임 매핑
SUBSTAT_COL_NAMES = {
    "speed": "속도",
    "atk": "공격력%",
    "hp": "생명력%",
    "def": "방어력%",
    "crit": "치명확률",
    "critdmg": "치명피해",
    "eff": "효과적중",
    "effres": "효과저항",
}

# 우선순위 문자열 약자 → 풀네임
# 치치 = 치확+치피 (특수 케이스)
PRIORITY_ABBREV = {
    "속": "속도",
    "공": "공격력%",
    "생": "생명력%",
    "방": "방어력%",
    "확": "치명확률",
    "피": "치명피해",
    "적": "효과적중",
    "저": "효과저항",
}

# 세트 약자 → 풀네임 (set keyword 컬럼 기반)
SET_ABBREV = {
    "속": "속도", "면": "면역", "생": "체력", "체": "체력",
    "격": "격류", "적": "적중", "관": "관통", "파": "파멸",
    "치": "치명", "저": "저항", "반": "반격", "방": "방어",
    "추": "추격", "흡": "흡혈", "역": "역습", "수": "수호",
    "개": "개전", "상": "상처", "응": "응수", "분": "분노",
    "공": "공격",
}

# 자유 텍스트 노트 신호 — 이거 만나면 그 alternate는 종료
NOTE_MARKERS = ["도 ", "도됨", "도 됨", "위에서", "추가", "가능"]


def parse_variant(name: str):
    """괄호 안 변형 표기 분리"""
    m = re.match(r"^(.+?)\s*\((.+)\)\s*$", name)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return name, None


def parse_marks(cell_val):
    """● → essential, ○ → preferred, 그 외 → None"""
    if cell_val is None:
        return None
    s = str(cell_val).strip()
    if s == "●":
        return "essential"
    if s == "○":
        return "preferred"
    return None


def parse_priority(s: str):
    """우선순위 문자열 → ordered list of full substat names"""
    if not s:
        return {"raw": "", "order": [], "unknown": False}
    s = s.strip()
    if "몰루" in s:
        return {"raw": s, "order": [], "unknown": True}

    order = []
    i = 0
    while i < len(s):
        ch = s[i]
        # 2자리 토큰 먼저 체크
        two = s[i:i+2]
        if two == "치확":
            order.append("치명확률"); i += 2; continue
        if two == "치피":
            order.append("치명피해"); i += 2; continue
        if two == "치치":
            order.append("치명확률"); order.append("치명피해"); i += 2; continue
        if ch == "치":
            # 단독 "치" — 보통 치명을 의미하나 모호. 뒤에 다른 글자가 오면 둘 다(치치) 가정.
            # 데이터셋 기준으로 단독 "치"는 거의 없음. 안전하게 둘 다 추가.
            order.append("치명확률"); order.append("치명피해"); i += 1; continue
        if ch in PRIORITY_ABBREV:
            order.append(PRIORITY_ABBREV[ch]); i += 1; continue
        # 알 수 없는 문자 — 스킵하고 보고
        i += 1
    return {"raw": s, "order": order, "unknown": False}


# 풀네임 세트 (긴 것부터 매칭하도록 정렬)
SET_FULL_NAMES = sorted(set(SET_ABBREV.values()) | {"생명"}, key=len, reverse=True)
# 생명 = 체력의 다른 표기 (생명력 set의 정식 이름)
SET_FULLNAME_NORMALIZE = {"생명": "체력"}


def parse_set_combo(s: str):
    """세트조합 문자열 → alternates list with metadata"""
    if not s:
        return {"raw": "", "alternates": [], "ignore_2set": False}
    s = s.strip()
    ignore_2set = False
    # "2셋무시" 플래그 처리
    if "2셋무시" in s:
        ignore_2set = True
        s_clean = s.replace("2셋무시", "")
    else:
        s_clean = s
    # / 로 분리 (대안)
    alts_raw = [a.strip() for a in s_clean.split("/")]
    alternates = []
    for alt in alts_raw:
        sets = []
        i = 0
        while i < len(alt):
            # 자유 텍스트 노트 시작이면 종료
            tail = alt[i:]
            if any(tail.startswith(m) for m in NOTE_MARKERS):
                break
            ch = alt[i]
            if ch in "+ ,":
                i += 1; continue
            # 풀네임 우선 매칭 (긴 것부터)
            matched = False
            for full in SET_FULL_NAMES:
                if alt[i:i+len(full)] == full:
                    canonical = SET_FULLNAME_NORMALIZE.get(full, full)
                    sets.append(canonical)
                    i += len(full)
                    matched = True
                    break
            if matched:
                continue
            # 약자 매칭
            if ch in SET_ABBREV:
                sets.append(SET_ABBREV[ch])
                i += 1
                continue
            # 알 수 없는 글자 — 스킵 (자유 텍스트 일부 가능성)
            i += 1
        if sets:
            # 중복 제거 (순서 유지)
            seen = set()
            uniq = []
            for x in sets:
                if x not in seen:
                    seen.add(x); uniq.append(x)
            alternates.append(uniq)
    return {"raw": s, "alternates": alternates, "ignore_2set": ignore_2set}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["영웅별 유효옵"]

    heroes = []
    report_lines = []
    seen_names = {}
    duplicates = []
    weird_priority = []
    weird_setcombo = []
    parens_count = 0

    for r in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=r, column=COL["name"]).value
        if not name_cell:
            continue
        name = str(name_cell).strip()
        if not name:
            continue

        base, variant = parse_variant(name)
        if variant:
            parens_count += 1

        # 부옵 마크
        valid_substats = {}
        for key, col in [("speed", 2), ("atk", 3), ("hp", 4), ("def", 5),
                         ("crit", 6), ("critdmg", 7), ("eff", 8), ("effres", 9)]:
            valid_substats[SUBSTAT_COL_NAMES[key]] = parse_marks(
                ws.cell(row=r, column=col).value
            )

        # 우선순위
        pri_val = ws.cell(row=r, column=COL["priority"]).value
        priority = parse_priority(str(pri_val).strip() if pri_val else "")
        # 우선순위에 풀어내지 못한 글자가 있는지 확인
        if pri_val:
            raw = str(pri_val).strip()
            recognized_chars = set("속공생방치확피적저") | set("몰루")
            unknown_chars = [c for c in raw if c not in recognized_chars]
            if unknown_chars:
                weird_priority.append((name, raw, unknown_chars))

        # 세트조합
        sc_val = ws.cell(row=r, column=COL["set_combo"]).value
        set_combo = parse_set_combo(str(sc_val).strip() if sc_val else "")
        # 자유 텍스트 노트가 섞인 경우 (alternates 파싱 결과가 valid_sets 보다 적게 잡힘)
        if sc_val and not set_combo["alternates"]:
            weird_setcombo.append((name, str(sc_val).strip(), "no alternates parsed"))

        # 세트 풀네임 키워드 (cols 21-33)
        valid_sets = []
        for c in SET_KW_COLS:
            v = ws.cell(row=r, column=c).value
            if v:
                valid_sets.append(str(v).strip())

        # 비고
        notes_val = ws.cell(row=r, column=COL["notes"]).value
        notes = str(notes_val).strip() if notes_val else None

        hero = {
            "name": name,
            "base_name": base,
            "variant": variant,
            "valid_substats": valid_substats,
            "priority": priority,
            "set_combo": set_combo,
            "valid_sets": valid_sets,
            "notes": notes,
        }
        heroes.append(hero)

        # 중복 체크
        if name in seen_names:
            duplicates.append((name, r, seen_names[name]))
        else:
            seen_names[name] = r

    # 저장
    DST_RAW.parent.mkdir(parents=True, exist_ok=True)
    with open(DST_RAW, "w", encoding="utf-8") as f:
        json.dump(heroes, f, ensure_ascii=False, indent=2)

    # 리포트
    report_lines.append(f"Total heroes: {len(heroes)}")
    report_lines.append(f"With variant (괄호 표기): {parens_count}")
    report_lines.append(f"Duplicates: {len(duplicates)}")
    for n, r, prev in duplicates:
        report_lines.append(f"  {n} at row {r} (also row {prev})")

    report_lines.append("")
    report_lines.append(f"Weird priority chars: {len(weird_priority)}")
    for n, raw, ucs in weird_priority:
        report_lines.append(f"  [{n}] '{raw}' unknown: {ucs}")

    report_lines.append("")
    report_lines.append(f"Set combos with unparsed chars: {len(weird_setcombo)}")
    for n, raw, uc in weird_setcombo[:30]:
        report_lines.append(f"  [{n}] '{raw}' unknown: '{uc}'")
    if len(weird_setcombo) > 30:
        report_lines.append(f"  ... and {len(weird_setcombo) - 30} more")

    # 부옵 마크는 있는데 우선순위에서 빠진 케이스
    report_lines.append("")
    report_lines.append("Sanity: substats marked but missing in priority order")
    mismatch = 0
    for h in heroes:
        if h["priority"]["unknown"]:
            continue
        marked = {k for k, v in h["valid_substats"].items() if v == "essential"}
        in_priority = set(h["priority"]["order"])
        missing = marked - in_priority
        if missing:
            mismatch += 1
            if mismatch <= 10:
                report_lines.append(f"  [{h['name']}] essential={marked}, priority={in_priority}, missing={missing}")
    report_lines.append(f"  ... total mismatches: {mismatch}")

    DST_REPORT.write_text("\n".join(report_lines), encoding="utf-8")
    print(f"Wrote {len(heroes)} heroes to {DST_RAW}")
    print(f"Report at {DST_REPORT}")
    print()
    print("=== Report preview ===")
    print("\n".join(report_lines[:60]))


if __name__ == "__main__":
    main()

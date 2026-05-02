import openpyxl
import sys
import io
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jsh02\Downloads\에픽세븐 장비 유효옵의 사본.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

# Deep analysis of 영웅별 유효옵
ws = wb["영웅별 유효옵"]
print("=" * 70)
print("영웅별 유효옵 — header row & sample rows")
print("=" * 70)

# Print header with column index
header = []
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=c).value
    header.append(f"col{c}({v})")
print("HEADER:", header)
print()

# Count actual hero rows (non-empty col A)
hero_count = 0
last_hero_row = 0
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None and str(v).strip():
        hero_count += 1
        last_hero_row = r
print(f"Total heroes: {hero_count}, last hero row: {last_hero_row}")
print()

# Sample some rows from middle and end
print("--- Mid sample (rows 100-103) ---")
for r in [100, 101, 102, 103]:
    row_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        row_vals.append(str(v) if v is not None else "")
    if any(row_vals):
        print(f"R{r}: {row_vals}")

print()
print("--- Last sample (rows 500-505) ---")
for r in [500, 501, 502, 503, 504, 505]:
    row_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        row_vals.append(str(v) if v is not None else "")
    if any(row_vals):
        print(f"R{r}: {row_vals}")

print()
print("--- 1015-1020 ---")
for r in range(1015, 1021):
    row_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        row_vals.append(str(v) if v is not None else "")
    if any(row_vals):
        print(f"R{r}: {row_vals}")

# Analyze unique values in 우선순위 column (col 10)
print()
print("--- 우선순위 (col 10) value variety ---")
priorities = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=10).value
    if v:
        priorities[str(v)] += 1
print(f"Total unique priority strings: {len(priorities)}")
print("Top 20:")
for k, c in priorities.most_common(20):
    print(f"  {c:>3} : {k}")

# Analyze 세트조합 column (col 17)
print()
print("--- 세트조합 (col 17) value variety ---")
sets = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=17).value
    if v:
        sets[str(v)] += 1
print(f"Total unique set combos: {len(sets)}")
print("Top 20:")
for k, c in sets.most_common(20):
    print(f"  {c:>3} : {k}")

# Look at columns 21+ (keyword columns?)
print()
print("--- Cols 21-33 sample ---")
for r in [2, 3, 4, 5, 6, 7, 8, 9, 10]:
    vals = []
    for c in range(20, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        vals.append(str(v) if v is not None else "")
    print(f"R{r} cols20-33: {vals}")

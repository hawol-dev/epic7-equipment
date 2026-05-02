import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jsh02\Downloads\에픽세븐장비시뮬_26_04_14의 사본.xlsx"
print("Loading large workbook (32MB)...")
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

def dump_sheet(name, max_rows=12):
    ws = wb[name]
    print("=" * 70)
    print(f"SHEET: {name}  (max_row={ws.max_row}, max_col={ws.max_column})")
    print("=" * 70)
    r = 0
    for row in ws.iter_rows(values_only=True):
        r += 1
        if r > max_rows:
            break
        vals = []
        for v in row:
            if v is None:
                vals.append("")
            else:
                s = str(v).replace("\n", " ")
                if len(s) > 25:
                    s = s[:25] + "..."
                vals.append(s)
        # only print non-empty
        if any(vals):
            print(f"R{r}: " + " | ".join(vals))
    print()

# Hit each sheet briefly
for s in wb.sheetnames:
    try:
        dump_sheet(s, max_rows=8)
    except Exception as e:
        print(f"Error on {s}: {e}")

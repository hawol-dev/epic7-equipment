"""
파일2 (장비시뮬.xlsx) → 작은 JSON 캐시 파일들로 추출.

CI 환경에는 xlsx 못 올리니 (32MB) 로컬에서 한 번 추출해서 JSON 으로 커밋.
build_unified_db.py 는 이 JSON 들을 읽어서 동작.

xlsx가 업데이트되면 이 스크립트 다시 돌려서 JSON 갱신.
"""
import json
import sys
import io
import openpyxl
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
SRC = Path(r"C:\Users\jsh02\Downloads\에픽세븐장비시뮬_26_04_14의 사본.xlsx")

DST_ARTIFACTS = ROOT / "data/raw/file2_artifacts.json"
DST_ENGRAVING = ROOT / "data/raw/file2_engraving_grades.json"
DST_META = ROOT / "data/raw/file2_meta.json"


def extract_artifacts(wb) -> dict[str, list[str]]:
    ws = wb["추천 세팅"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        if name == "영웅 선택":
            continue
        artifacts = [row[2], row[3], row[4]]
        artifacts = [str(a).strip() for a in artifacts if a]
        if artifacts:
            out[name] = artifacts
    return out


def extract_engraving_grades(wb) -> dict:
    GRADES = ["D", "C", "B", "A", "S", "SS", "SSS"]
    # ENGRAVING_KO_TO_ID — enums.py 와 동일하게 (의존성 피해 inline)
    ENGRAVING_KO_TO_ID = {
        "공격력(%)": "atk_p", "방어력(%)": "def_p", "생명력(%)": "hp_p",
        "치명확률": "chc", "치명피해": "chd",
        "효과적중": "eff", "효과저항": "effres",
        "공격력": "atk", "방어력": "def", "생명력": "hp",
    }
    ws = wb["계산요소"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        ko_stat = row[0]
        level = row[1]
        if not ko_stat or level is None:
            continue
        ko_stat = str(ko_stat).strip()
        try:
            level_int = int(level)
        except (TypeError, ValueError):
            continue
        sid = ENGRAVING_KO_TO_ID.get(ko_stat)
        if not sid:
            continue
        grade_values = {}
        for gi, grade in enumerate(GRADES):
            v = row[2 + gi]
            grade_values[grade] = str(v).strip() if v is not None else None
        out.setdefault(sid, {})[level_int] = grade_values
    return out


def extract_meta(wb) -> dict:
    """캐릭터 정보 시트 → 영웅 메타. heroes_merged.json 머지용 백업."""
    ws = wb["캐릭터 정보"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        if not name or str(name).strip() == "영웅 선택":
            continue
        name = str(name).strip()
        out[name] = {
            "rank": int(row[2]) if row[2] else None,
            "element": str(row[3]).strip() if row[3] else None,
            "zodiac": str(row[4]).strip() if row[4] else None,
            "class": str(row[5]).strip() if row[5] else None,
            "engraving_focus": str(row[6]).strip() if row[6] else None,
        }
    return out


def main():
    if not SRC.exists():
        print(f"❌ xlsx 없음: {SRC}")
        sys.exit(1)

    print(f"Loading xlsx ({SRC.stat().st_size // 1024 // 1024}MB)...")
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    artifacts = extract_artifacts(wb)
    engravings = extract_engraving_grades(wb)
    meta = extract_meta(wb)

    DST_ARTIFACTS.write_text(
        json.dumps(artifacts, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    DST_ENGRAVING.write_text(
        json.dumps({str(sid): {str(k): v for k, v in lv.items()}
                    for sid, lv in engravings.items()},
                   ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    DST_META.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"✓ 추천 아티팩트: {len(artifacts)}명 → {DST_ARTIFACTS.name}")
    print(f"✓ 각인집중 등급표: {len(engravings)} 스탯 → {DST_ENGRAVING.name}")
    print(f"✓ 캐릭터 메타: {len(meta)}명 → {DST_META.name}")
    wb.close()


if __name__ == "__main__":
    main()

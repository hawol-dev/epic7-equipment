"""파일2 캐릭터 정보 탭 전체 검사"""
import openpyxl
import sys
import io
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jsh02\Downloads\에픽세븐장비시뮬_26_04_14의 사본.xlsx"
print("Loading...")
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb["캐릭터 정보"]

# 헤더 + 처음 5행
header = None
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        header = list(row)
        continue
    rows.append(list(row))
    if i >= 500:  # 어디까지 데이터가 있는지 보자
        break

print("HEADER (cols 1-15):")
for ci, c in enumerate(header[:15], 1):
    print(f"  col{ci}: {c}")

# 비어있지 않은 행만
non_empty = [r for r in rows if r[0] and str(r[0]).strip() and str(r[0]).strip() != "영웅 선택"]
print(f"\nNon-empty hero rows in first 500: {len(non_empty)}")

# 특수값 확인
print(f"\nFirst 5 non-empty rows:")
for r in non_empty[:5]:
    print(f"  {r[:14]}")

# 등급, 속성, 직업, 별자리 unique values
ranks = Counter()
elements = Counter()
classes = Counter()
zodiacs = Counter()
engravings = Counter()

for r in non_empty:
    if r[2]: ranks[str(r[2])] += 1
    if r[3]: elements[str(r[3])] += 1
    if r[4]: zodiacs[str(r[4])] += 1
    if r[5]: classes[str(r[5])] += 1
    if r[6]: engravings[str(r[6])] += 1

print(f"\n등급: {dict(ranks)}")
print(f"속성: {dict(elements)}")
print(f"직업: {dict(classes)}")
print(f"별자리 ({len(zodiacs)}종): {dict(zodiacs)}")
print(f"각인집중 ({len(engravings)}종): {dict(engravings)}")

# 끝까지 스캔해서 전체 영웅 수 확인
print("\n전체 행 스캔 중...")
all_heroes = set()
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    name = row[0]
    if name and str(name).strip() and str(name).strip() != "영웅 선택":
        all_heroes.add(str(name).strip())
print(f"파일2 캐릭터 정보 탭 영웅 수: {len(all_heroes)}")

"""우선순위/세트조합 문자열 정확히 분석 (수정: col18=세트조합, col19=비고)"""
import openpyxl
import sys
import io
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jsh02\Downloads\에픽세븐 장비 유효옵의 사본.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["영웅별 유효옵"]

PRIORITY_COL = 10
SETCOMBO_COL = 18
NOTES_COL = 19
SET_KEYWORD_START_COL = 21  # actually 24+ has data, but scan 21-33

priority_chars = Counter()
setcombo_chars = Counter()
set_keywords = Counter()
weird_priority_heroes = {}  # token -> [hero names]
weird_setcombo_heroes = {}

for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if not name:
        continue
    name = str(name).strip()

    pri = ws.cell(row=r, column=PRIORITY_COL).value
    if pri:
        s = str(pri).strip()
        for ch in s:
            priority_chars[ch] += 1

    setc = ws.cell(row=r, column=SETCOMBO_COL).value
    if setc:
        s = str(setc).strip()
        for ch in s:
            setcombo_chars[ch] += 1
            if ch in '/+,()&-':
                weird_setcombo_heroes.setdefault(ch, []).append((name, s))

    for c in range(SET_KEYWORD_START_COL, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v:
            set_keywords[str(v).strip()] += 1

print("=== Priority col10 char frequency ===")
for ch, c in priority_chars.most_common():
    print(f"  '{ch}' x{c}")

print()
print("=== Set combo col18 char frequency ===")
for ch, c in setcombo_chars.most_common():
    print(f"  '{ch}' x{c}")

print()
print("=== Set keyword cols 21-33 (full names) ===")
for ch, c in set_keywords.most_common():
    print(f"  '{ch}' x{c}")

print()
print("=== Set combos with special chars ===")
for ch, hits in weird_setcombo_heroes.items():
    print(f"--- '{ch}' ({len(hits)} heroes) ---")
    for name, s in hits[:5]:
        print(f"  [{name}] {s}")

# Find heroes with weird priority chars
print()
print("=== Heroes with weird priority chars (몰, 루, 피, 확) ===")
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    pri = ws.cell(row=r, column=PRIORITY_COL).value
    if pri and any(ch in str(pri) for ch in '몰루피확'):
        print(f"  [{name}] 우선순위='{pri}'")

# Cross-check substat marks vs priority string
print()
print("=== First 20 heroes — substat marks vs priority ===")
print("name | 속 공 생 방 치확 치피 효적 효저 | 우선순위 | 세트조합")
for r in range(2, 22):
    name = ws.cell(row=r, column=1).value
    if not name:
        continue
    marks = [ws.cell(row=r, column=c).value or "" for c in range(2, 10)]
    pri = ws.cell(row=r, column=10).value or ""
    setc = ws.cell(row=r, column=18).value or ""
    setkw = [ws.cell(row=r, column=c).value or "" for c in range(21, 34)]
    setkw_str = "/".join([s for s in setkw if s])
    print(f"  {name} | {' '.join(marks)} | {pri} | {setc} | sets={setkw_str}")
